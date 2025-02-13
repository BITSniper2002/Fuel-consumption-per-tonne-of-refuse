import tkinter as tk
import csv
import os
import openpyxl
import matplotlib.pyplot as plt
import numpy as np

root = tk.Tk()
root.geometry("300x200")
root.title('融江金顺车辆每次加油的百公里油耗')
root.configure(bg='white')
# 创建 Label、Entry 和 Button 对象
label = tk.Label(root, text="请输入文件名:",font=('微软雅黑', 10, 'bold'))
label.pack(pady=10)
entry = tk.Entry(root, width=50,background='red')
entry.pack()
button = tk.Button(root, text="运行",font=('微软雅黑', 10, 'bold'))

# 定义打开CSV文件的函数
def open_csv_file():
    # 打开 Excel 文件
    # 获取输入的文件名
    file_name = entry.get()
    # 拼接文件路径
    workbook = openpyxl.load_workbook(file_name)

    # 选择工作表
    worksheet = workbook.active

    # 处理每一个工作表
    for sheet_name in workbook.sheetnames:
            # print(sheet_name)
        if sheet_name == '加油明细':
            worksheet = workbook[sheet_name]
            cnt = 0
            valid_data = []
            for row in worksheet.iter_rows(values_only=True):
                # 选择有用的列输出
                new_row = row[1:5] + row[6:8]
                # 处理每一行数据
                if new_row[0] != None:
                    # print(new_row)
                    valid_data.append(new_row)
                    cnt += 1

    # 输出到valid.csv文件中
    with open('valid_test.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for row in valid_data:
            # 如果公里数、加油量、加油金额没有空缺，说明这条数据有效，输出
            if row[3] and row[4] and row[5]:
                writer.writerow(row)

    # 列表用于存储每一辆不同的车
    distinct_car = []

    # 打开前面处理完的文件
    with open('valid_test.csv', 'r', newline='') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            # 寻找独特的车牌号
            if row[2] not in distinct_car and row[2] != '车牌号':
                distinct_car.append(row[2])

    # 打开前面处理完的文件
    with open('valid_test.csv', 'r', newline='') as csvfile:
        reader = csvfile.readlines()
        # 读取标题
        title = reader[0]
        if not os.path.exists('./test_cars'):
            os.mkdir('./test_cars')
        # 将每一辆不同的车的相关信息都输出到其相应的csv文件（在cars文件夹下）
        while distinct_car:
            tmp = distinct_car.pop(0)
            data = []
            for row in reader:
                if tmp in row:
                    data.append(row)
            # print(data)
            with open('./test_cars/' + tmp + '.csv', 'w', newline='') as new_csvfile:
                new_csvfile.writelines(title)
                new_csvfile.writelines(data)
    folder_path = './test_cars'

    # 遍历文件夹下的所有文件
    for filename in os.listdir(folder_path):
        # 拼接文件路径
        filepath = os.path.join(folder_path, filename)
        # 打开存储每辆车信息的独特的csv文件
        with open(filepath, 'r', newline='') as csvfile:
            reader = csvfile.readlines()
            title = reader[0]
            data = []
            # 将数据从字符串转成列表类型
            for row in reader[1:]:
                tmp = row.split(',')
                tmp[4], tmp[5] = round(float(tmp[4]), 2), round(float(tmp[5]), 2)
                data.append(tmp)
            if len(data):
                new_data = [data[0]]
            else:
                break
            for i in range(1, len(data)):
                # 判断第四个元素是否相同
                if data[i][3] == new_data[-1][3]:
                    # 合并第五个元素和第六个元素的值
                    new_data[-1][4] += data[i][4]
                    new_data[-1][5] += data[i][5]
                else:
                    # 将不同的数组添加到新的列表中
                    new_data.append(data[i])
            w = []
            # 将列表转回字符串用于输出
            for item in new_data:
                x = ','.join(str(i) for i in item) + '\n'
                w.append(x)

            # 根据车牌号输出
            with open(filepath, 'w', newline='') as new_csvfile:
                new_csvfile.writelines(title)
                new_csvfile.writelines(w)
    folder_path = './test_cars'
    if not os.path.exists('./test_result'):
        os.mkdir('./test_result')

    # 如果total文件存在，删除该文件
    if os.path.isfile('./total.txt'):
        os.remove('./total.txt')

    # 遍历文件夹下的所有文件
    for filename in os.listdir(folder_path):
        # 拼接文件路径
        filepath = os.path.join(folder_path, filename)

        # 打开存储不同车辆信息的文件
        with open(filepath, 'r', newline='') as csvfile:
            reader = csvfile.readlines()
            title = reader[0]
            data = []
            consumption = []

            # 将字符串转成列表形式
            for row in reader[1:]:
                tmp = row.split(',')
                tmp[4], tmp[5] = round(float(tmp[4]), 2), round(float(tmp[5]), 2)
                data.append(tmp)
            first = data[0]

            # 求百公里油耗
            for i in range(1, len(data)):
                kilo_dif = float(data[i][3]) - float(data[i - 1][3])
                per = float(data[i - 1][4]) / kilo_dif * 100
                consumption.append(per)

            # 避免下一行出现除以0的情况
            if len(consumption):
                mean_value = sum(consumption) / len(consumption)
                variance = np.var(consumption)
                max_value, min_value = max(consumption), min(consumption)

                # 打开total文件，选择在文件后面添加内容的写入方式，写入最大值、最小值、平均值、方差
                with open('./total.txt', 'a') as file:
                    file.write(filename.rstrip('.csv') + ':')
                    file.write('最大值：' + str(round(max_value, 2)) + '最小值：' + str(round(min_value, 2)) +
                               '平均值：' + str(round(mean_value, 2)) + '方差：' + str(round(variance, 2)) + '\n')
                # print(filepath,':',max_value,min_value,mean_value,variance)

                # 定义异常范围为超过平均值百分之十
                max_range = mean_value * 1.1

                # 找到异常值的序号
                abnormal = [i for i, num in enumerate(consumption) if num > max_range]
                date = []

                # 以时间，百公里耗油的形式输出到result文件
                for i in abnormal:
                    date.append(str(data[i][0]) + ',' + str(consumption[i]) + "\n")
                with open('./test_result/_' + filename, 'w') as file:
                    file.write('时间,百公里耗油\n')
                    file.writelines(date)
                if not os.path.exists('./test_pics'):
                    os.mkdir('./test_pics')
                # 为每辆车的结果建立新图像
                plt.figure()
                x = range(len(consumption))
                y = consumption
                plt.plot(x, y)
                # 添加平均值线
                plt.axhline(y=mean_value, color='r', linestyle='--')
                plt.axhline(y=max_range, color='black', linestyle='--')
                # 添加标签和标题
                plt.xlabel('第几次加油')
                plt.ylabel('百公里耗油')
                plt.title('Line Plot')
                # 显示图形
                plt.savefig('./test_pics/' + filename + '.png')
            else:
                continue



# 绑定按钮点击事件
button.config(command=open_csv_file)
button.pack(pady=10)

root.mainloop()